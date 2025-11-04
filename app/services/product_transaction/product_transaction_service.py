from ...models.product_transaction.product_transaction import ProductTransaction
from ...services.transaction_type.transaction_type_service import TransactionTypeService
from ...services.inventory.inventory_service import InventoryService
from ...services.branch.branch_service import BranchService
from ...services.product.product_service import ProductService
from ...services.supplier.supplier_service import SupplierService
from ...services.log.log_service import LogService
from ...services.staff.staff import get_user_by_id
from ...database import db
from ...utils.validator import validate_data
from ...utils.date_conversor import parse_transaction_date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


class ProductTransactionService:

    @staticmethod
    def get_all_products_transactions():

        product_transactions = ProductTransaction.query.all()

        return [p_transaction.to_dict() for p_transaction in product_transactions]

    @staticmethod
    def get_product_transaction_by_id(id_product_transaction):

        product_transaction = ProductTransaction.query.filter(
            ProductTransaction.id == id_product_transaction
        ).first()

        if product_transaction is None:
            LogService.create_log(
                {
                    "module": f"{ProductTransactionService.__name__}.{ProductTransactionService.get_product_transaction_by_id.__name__}",
                    "message": "No se encontró la transacción buscada por id",
                }
            )
            raise ValueError("Transacción no encontrada")

        return product_transaction.to_dict()

    @staticmethod
    def create_product_transaction_service(product_transaction):

        required_fields = {
            "description": str,
            "quantity": int,
            "unit_price": (int, float),
            "transaction_date": str,
            "product_id": (int, str),
            "branch_id": (int, str),
            "transaction_type_id": (int, str),
            "app_user_id": (int, str),
        }

        validate_data(product_transaction, required_fields)

        ProductTransactionService.validate_product_transaction_data(product_transaction)

        unit_price = Decimal(product_transaction["unit_price"])
        total_price = unit_price * product_transaction["quantity"]

        # Validar y parsear la fecha con formatos estrictos
        parsed_date = parse_transaction_date(product_transaction["transaction_date"])

        transaction_type = TransactionTypeService.get_transaction_type_by_id(
            product_transaction["transaction_type_id"]
        )

        try:

            InventoryService.update_inventory(product_transaction, transaction_type)

            # Crear la transacción de producto
            new_product_transaction = ProductTransaction(
                description=product_transaction["description"],
                quantity=product_transaction["quantity"],
                unit_price=unit_price,
                total_price=total_price,
                transaction_date=parsed_date,
                product_id=product_transaction["product_id"],
                supplier_id=product_transaction.get("supplier_id"),  # Opcional
                branch_id=product_transaction["branch_id"],
                transaction_type_id=product_transaction["transaction_type_id"],
                app_user_id=product_transaction["app_user_id"],
            )

            db.session.add(new_product_transaction)
            db.session.commit()

            return new_product_transaction

        except Exception as e:
            db.session.rollback()
            LogService.create_log(
                {
                    "module": f"{ProductTransactionService.__name__}.{ProductTransactionService.create_product_transaction_service.__name__}",
                    "message": f"Ocurrió un error en la creación de la transacción del producto: {str(e)}. Se realizó rollback.",
                }
            )
            raise e

    @staticmethod
    def validate_product_transaction_data(product_transaction):

        ProductService.get_product_by_id(product_transaction["product_id"])

        BranchService.get_branch_by_id(product_transaction["branch_id"])

        get_user_by_id(
            product_transaction["app_user_id"]
        )  # Validamos si el usuario de la transacción existe

        if "supplier_id" in product_transaction and product_transaction["supplier_id"]:
            SupplierService.get_supplier_by_id(product_transaction["supplier_id"])

        if product_transaction["quantity"] < 0:
            LogService.create_log(
                {
                    "module": f"{ProductTransactionService.__name__}.{ProductTransactionService.validate_product_transaction_data.__name__}",
                    "message": "Se ingresó una cantidad negativa en la transacción",
                }
            )
            raise ValueError("La Cantidad no puede ser negativa")

        if len(product_transaction["description"].strip()) < 5:
            LogService.create_log(
                {
                    "module": f"{ProductTransactionService.__name__}.{ProductTransactionService.validate_product_transaction_data.__name__}",
                    "message": "Se ingresó una descripción menor a 5 caracterese en la transacción",
                }
            )
            raise ValueError("La descripción debe ser mayor a 5 caracteres")

        if product_transaction["unit_price"] < 0:
            LogService.create_log(
                {
                    "module": f"{ProductTransactionService.__name__}.{ProductTransactionService.validate_product_transaction_data.__name__}",
                    "message": "Se ingresó una precio unitario negativo en la transacción",
                }
            )
            raise ValueError(f"El precio unitario no puede ser negativo")

    @staticmethod
    def generate_excel_report():
        """
        Genera un archivo Excel con todas las transacciones de productos
        """
        try:
            # Obtener todas las transacciones
            product_transactions = ProductTransaction.query.all()

            # Crear el workbook y la hoja activa
            wb = Workbook()
            ws = wb.active
            ws.title = "Transacciones de Productos"

            # Definir estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Definir encabezados
            headers = [
                "ID",
                "Descripción",
                "Producto",
                "Cantidad",
                "Precio Unitario",
                "Precio Total",
                "Fecha de Transacción",
                "Tipo de Transacción",
                "Sucursal",
                "Usuario",
                "Proveedor",
                "Fecha de Creación"
            ]

            # Escribir encabezados
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Escribir datos
            for row_num, transaction in enumerate(product_transactions, 2):
                ws.cell(row=row_num, column=1, value=transaction.id)
                ws.cell(row=row_num, column=2, value=transaction.description)
                ws.cell(row=row_num, column=3, value=transaction.product.name if transaction.product else "N/A")
                ws.cell(row=row_num, column=4, value=transaction.quantity)
                ws.cell(row=row_num, column=5, value=float(transaction.unit_price))
                ws.cell(row=row_num, column=6, value=float(transaction.total_price))
                ws.cell(row=row_num, column=7, value=transaction.transaction_date.strftime("%Y-%m-%d %H:%M:%S") if transaction.transaction_date else "N/A")
                ws.cell(row=row_num, column=8, value=transaction.transaction_type.name if transaction.transaction_type else "N/A")
                ws.cell(row=row_num, column=9, value=transaction.branch.name if transaction.branch else "N/A")
                ws.cell(row=row_num, column=10, value=transaction.app_user.name if transaction.app_user else "N/A")
                ws.cell(row=row_num, column=11, value=transaction.supplier.name if transaction.supplier else "N/A")
                ws.cell(row=row_num, column=12, value=transaction.created_at.strftime("%Y-%m-%d %H:%M:%S") if transaction.created_at else "N/A")

            # Ajustar el ancho de las columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar en un BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            return excel_file

        except Exception as e:
            LogService.create_log(
                {
                    "module": f"{ProductTransactionService.__name__}.{ProductTransactionService.generate_excel_report.__name__}",
                    "message": f"Error al generar el reporte Excel: {str(e)}",
                }
            )
            raise e
